import os
from openpyxl import Workbook
import openpyxl

ros = 0
col = 0
files =0
filelist = os.listdir('excel files')
directory = 'excel files'
number_of_files = len([item for item in os.listdir(directory) if os.path.isfile(os.path.join(directory, item))])
print(number_of_files)

wb2 = openpyxl.load_workbook('excel files/'+filelist[0])
ws2 = wb2.active

row_count = ws2.max_row
column_count = ws2.max_column
newmatrix = [[0]*column_count for i in range(row_count)]
matrix = [[[0]*column_count for i in range(row_count)] for x in range(number_of_files)] 

for file in filelist:
    if file[0] == '~':continue
    wb = openpyxl.load_workbook('excel files/'+file)
    ws = wb.active    

    for row in ws.rows:
        for cell in row:
            #print(cell.value)
            matrix[files][ros][col] = cell.value
            col=col+1
        col = 0    
        ros=ros+1
    ros=0
    files=files+1

        
print( '\n\npress enter to exit')
#print(matrix)
a=0
b=0
c=0
for a in range(0,number_of_files):
    for b in range(0,row_count):
        for c in range(0,column_count):
            if type(matrix[a][b][c]) != int : matrix[a][b][c] =0
            newmatrix[b][c]=newmatrix[b][c]+ matrix[a][b][c]
print(newmatrix) 

savewb = Workbook()
ws3 = savewb.active # 현재 열려 있는 Sheet
ws3.title = "result"  #시트 ws1의 이름을 지정

for b in range(1,row_count+1):
    for c in range(1,column_count+1):
        ws3.cell(row=b,column=c, value= newmatrix[b-1][c-1])
 
savewb.save('0316.xlsx') #저장


